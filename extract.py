#!/usr/bin/env python3
"""
extract.py — pull NIFTY 50 (or a subset) from indianapi.in -> one Excel workbook.

Logging: every run writes a timestamped log to ./logs/ AND prints a per-stock
status summary at the end, so empty/failed stocks are visible (with the reason
the API gave) and you can take it to the API team.

Quota: cached stocks cost 0 calls; empty/failed responses are NOT cached, so
they retry on the next run. Run prints calls spent vs your monthly budget.

Run:
  export INDIANAPI_KEY=your-key
  python extract.py                 # stocks listed in config.NIFTY_50
  python extract.py --only TCS,SUNPHARMA --refresh
"""

import argparse
import json
import sys
import datetime
from pathlib import Path

import pandas as pd

import config
import client
import parse

STMT_TABS = ["Income", "Balance", "Cash Flow"]
LOG_DIR = config.ROOT / "logs"


def _write_excel(path, overview, stmts, metrics, holding, analysis_df, flags_df, meta):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        pd.DataFrame(list(meta.items()), columns=["field", "value"]).to_excel(
            xl, sheet_name="ReadMe", index=False)
        overview.to_excel(xl, sheet_name="Overview", index=False)
        analysis_df.to_excel(xl, sheet_name="Analysis", index=False)
        flags_df.to_excel(xl, sheet_name="Flags", index=False)
        for tab in STMT_TABS:
            (stmts[tab] if not stmts[tab].empty
             else pd.DataFrame({"info": ["(no data)"]})).to_excel(xl, sheet_name=tab, index=False)
        metrics.to_excel(xl, sheet_name="KeyMetrics", index=False)
        holding.to_excel(xl, sheet_name="Shareholding", index=False)
        for ws in xl.sheets.values():
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
            for i, _ in enumerate(ws[1], start=1):
                ws.column_dimensions[get_column_letter(i)].width = 24 if i <= 3 else 14


def main():
    ap = argparse.ArgumentParser(description="indianapi.in NIFTY 50 extractor")
    ap.add_argument("--refresh", action="store_true", help="ignore cache, re-fetch")
    ap.add_argument("--only", default=None, help="comma list of tickers (subset)")
    args = ap.parse_args()

    LOG_DIR.mkdir(exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    logfile = LOG_DIR / f"extract_{stamp}.log"
    client.configure_logging(logfile)
    log = client.log

    universe = dict(config.NIFTY_50)
    if args.only:
        want = {s.strip().upper() for s in args.only.split(",")}
        universe = {k: v for k, v in universe.items() if k in want}

    try:
        client.api_key()
    except RuntimeError as e:
        log.error(str(e)); sys.exit(1)

    log.info("Run start: %d stocks, refresh=%s, endpoint=%s",
             len(universe), args.refresh, config.API_URL)

    raw_all, parsed, summary = {}, [], []
    for i, (tk, name) in enumerate(universe.items(), 1):
        before = client.calls_made()
        try:
            raw = client.get_stock(tk, name, refresh=args.refresh)
        except Exception as e:  # noqa: BLE001  (network/other)
            log.error("%-12s EXCEPTION: %s", tk, e)
            summary.append((tk, name, "ERROR", str(e)[:80]))
            continue

        spent = client.calls_made() > before
        if raw is None:
            summary.append((tk, name, "EMPTY", "see log (API returned no usable data)"))
            continue
        if not client.is_valid(raw):
            summary.append((tk, name, "EMPTY", "cached but invalid"))
            continue

        raw_all[tk] = raw
        stmts, _ = parse.statements(raw)
        parsed.append((tk, raw.get("companyName") or name, raw, stmts))
        summary.append((tk, name, "OK" + (" (api)" if spent else " (cache)"),
                        raw.get("companyName")))

    log.info("API calls spent this run: %d (of %d/month)",
             client.calls_made(), config.MONTHLY_BUDGET)

    # ---- per-stock status summary ----
    log.info("=" * 64)
    log.info("%-12s %-18s %s", "TICKER", "STATUS", "DETAIL")
    log.info("-" * 64)
    for tk, name, status, detail in summary:
        lvl = log.info if status.startswith("OK") else log.warning
        lvl("%-12s %-18s %s", tk, status, detail)
    log.info("=" * 64)
    ok = sum(1 for _, _, s, _ in summary if s.startswith("OK"))
    bad = [tk for tk, _, s, _ in summary if not s.startswith("OK")]
    log.info("OK: %d   |   EMPTY/ERROR: %d  %s", ok, len(bad),
             ("-> " + ", ".join(bad)) if bad else "")

    if not parsed:
        log.error("No usable data. Check the log above for the API responses.")
        sys.exit(1)

    # ---- assemble tables ----
    overview = pd.DataFrame([parse.snapshot(raw, tk) for tk, _, raw, _ in parsed])
    stmt_tables = {}
    for tab in STMT_TABS:
        frames = []
        for tk, comp, _, stmts in parsed:
            df = stmts.get(tab)
            if df is None or df.empty:
                continue
            d = df.reset_index()
            d.insert(0, "Ticker", tk)
            d.insert(1, "Company", comp)
            frames.append(d)
        stmt_tables[tab] = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    metrics = pd.concat([parse.metrics_long(raw, tk) for tk, _, raw, _ in parsed], ignore_index=True)
    holding = pd.concat([parse.shareholding(raw, tk) for tk, _, raw, _ in parsed], ignore_index=True)

    # multi-year analysis (runs on cached data; no API calls)
    import analysis
    analysis_df = pd.DataFrame([analysis.summary_row(raw, tk) for tk, _, raw, _ in parsed])
    flag_rows = []
    for tk, _, raw, _ in parsed:
        flag_rows.extend(analysis.flags_long(raw, tk))
    flags_df = pd.DataFrame(flag_rows) if flag_rows else pd.DataFrame(
        columns=["ticker", "type", "flag", "note"])
    n_invest = int(analysis_df["investigate"].sum()) if not analysis_df.empty else 0
    log.info("Analysis: %d/%d stocks flagged 'investigate' (>=%d yrs, no severe red flags)",
             n_invest, len(analysis_df), analysis.MIN_YEARS)

    meta = {
        "source": "indianapi.in (stock endpoint)",
        "generated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "stocks_ok": len(parsed),
        "stocks_empty_or_error": len(bad),
        "empty_or_error_tickers": ", ".join(bad) if bad else "(none)",
        "log_file": logfile.name,
        "note": "All money values in Rs crore. Statement years are columns.",
    }

    xlsx = config.OUTPUT_DIR / "nifty50_fundamentals.xlsx"
    _write_excel(xlsx, overview, stmt_tables, metrics, holding, analysis_df, flags_df, meta)
    with open(config.OUTPUT_DIR / "nifty50_data.json", "w", encoding="utf-8") as f:
        json.dump(raw_all, f)

    log.info("Wrote %s", xlsx)
    log.info("Wrote %s (for the app)", config.OUTPUT_DIR / "nifty50_data.json")
    log.info("Full log: %s", logfile)


if __name__ == "__main__":
    main()
